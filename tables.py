#! /usr/bin/env python
#
#                    _...
#              o_.-"`    `\				#############################################
#       .--.  _ `'-._.-'""-;     _ 		#############################################
#     .'    \`_\_  {_.-a"a-}  _ / \		##											#
#   _/     .-'  '. {c-._o_.){\|`  |		##			Coded by Avery Uslaner			#
#  (@`-._ /       \{    ^  } \\ _/		##				12/10/2015					#
#   `~\  '-._      /'.     }  \}  .-. 	##											#	
#     |>:<   '-.__/   '._,} \_/  / ())  ##											#
#     |     >:<   `'---. ____'-.|(`"`	##											#
#     \            >:<  \\_\\_\ | ;		##				#############				#
#      \                 \\-{}-\/  \	##				#			#				#
#       \                 '._\\'   /)	##				#			#				#
#        '.                       /(	##				#			#				#
#          `-._ _____ _ _____ __.'\ \	##				#			#				#
#            / \     / \     / \   \ \ 	##				#		 #  #				#
#     jgs _.'/^\'._.'/^\'._.'/^\'.__) \	##				#			#				#
#     ,=='  `---`   '---'   '---'      )##				#			#				#
#     `"""""""""""""""""""""""""""""""`	##				#			#				#
#####################################################################################

import openpyxl
import numpy
from openpyxl.styles import Color, PatternFill
import pdb
from math import fsum

wb_survivorship = openpyxl.load_workbook('2013-2015 Survivorship Table.xlsx')
wb_Fall_2013 = openpyxl.load_workbook('Seedling Data_Fall 2013_COPY.xlsx')
wb_Spring_2014 = openpyxl.load_workbook('Seedling Data Spring_April 2014_COPY.xlsx')
wb_Fall_2014 = openpyxl.load_workbook('Seedling Data Fall 2014.xlsx')
wb_Spring_2015 = openpyxl.load_workbook('Seedling Data Sp 2015_COPY.xlsx')
wb_Fall_2015 = openpyxl.load_workbook('Seedling Data Fall 2015_COPY.xlsx')

# I'm abusing global variables here but fuck it
cohort1_seedling = 0
cohort1_avg_life = []
cohort2_seedling = 0
cohort2_avg_life = []
cohort3_seedling = 0
cohort3_avg_life = []
cohort4_seedling = 0
cohort4_avg_life = []
no_cohort_seedling = 0

site1_seedling = 0
site1_avg_life = []
site2_seedling = 0
site2_avg_life = []
site3_seedling = 0
site3_avg_life = []
site4_seedling = 0
site4_avg_life = []

aspectN_seedling = 0
aspectN_avg_life = []
aspectS_seedling = 0
aspectS_avg_life = []

slopeA_seedling = 0
slopeA_avg_life = []
slopeB_seedling = 0
slopeB_avg_life = []
slopeC_seedling = 0
slopeC_avg_life = []

total_plants = 0
never_seedling = 0 # When we first found it, it didn't meet the criteria for seedling

workbooks = {
			'Fall2013': wb_Fall_2013,
			'Spring2014': wb_Spring_2014,
			'Fall2014': wb_Fall_2014,
			'Spring2015': wb_Spring_2015,
			'Fall2015': wb_Fall_2015
			}

cohort1Color = PatternFill(start_color='FF6666',
							end_color='FF6666',
							fill_type='solid')

cohort2Color = PatternFill(start_color='99CC66',
							end_color='99CC66',
							fill_type='solid')

cohort3Color = PatternFill(start_color='00bfff',
							end_color='00bfff',
							fill_type='solid')

cohort4Color = PatternFill(start_color='bfff00',
							end_color='bfff00',
							fill_type='solid')

#####################################################################
# A Class of plant objects to store relevant info
#####################################################################
class Individual(object):
	plants = []

	def __init__(self, site, x_y, aspect, slope, accession, year, condition, height, width, leaves, seedling):
		self.site = site
		self.x_y = x_y
		self.aspect = aspect
		self.slope = slope
		self.accession = accession
		self.years = [year]
		self.condition = condition
		self.height = height
		self.width = width
		self.leaves = leaves

		# Starting seedling values
		self.seedling = {
					'Fall 2013': False,
					'Spring 2014': False,
					'Fall 2014': False,
					'Spring 2015': False,
					'Fall 2015': False
		}

		# If passed in seedling value isn't None, update seedling dict
		if seedling != None:
			self.seedling[seedling] = True

		self.cohort = 99
		Individual.plants.append(self)

	def __str__(self):
		return "Site: {}\nX: {}\nY: {}\nAspect: {}\nSlope: {}\nAccession: {}\nYears: {}\nCondition: {}\nHeight: {}\nWidth: {}\nLeaves: {}\nSeedling Years: {}\n\n".format(self.site, self.x_y[0], self.x_y[1], self.aspect, self.slope, self.accession, self.years, self.condition, self.height, self.width, self.leaves, self.seedling)

	def __repr__(self):
		return "Site: {}\nX: {0}\nY: {1}\nAspect: {2}\nSlope: {3}\nAccession: {4}\nYears: {5}\nCondition: {6}\nHeight: {7}\nWidth: {8}\nLeaves: {8}\nSeedling Years: {9}\n\n".format(self.site, self.x_y[0], self.x_y[1], self.aspect, self.slope, self.accession, self.years, self.condition, self.height, self.width, self.leaves, self.seedling)	


	def addYear(self, year):
		self.years.append(year)

	def addSeedlingYear(self, seedling):
		self.seedling[seedling] = True

	def addCondition(self, condition):
		self.condition.update(condition)

	def addHeight(self, height):
		self.height.update(height)

	def addWidth(self, width):
		self.width.update(width)

	def addLeaves(self, leaves):
		self.leaves.update(leaves)

	def updateCohort(self, cohort):
		self.cohort = cohort

#####################################################################
# Figure out colors for the excel sheet
#####################################################################
def estColor(plant):
	if plant.cohort == 1:
		return cohort1Color
	elif plant.cohort == 2:
		return cohort2Color
	elif plant.cohort == 3:
		return cohort3Color
	elif plant.cohort == 4:
		return cohort4Color
	else:
		return None

#####################################################################
# Once all relevant data has been added to plant objects
# You can pass them in to check which cohort they should belong to
#####################################################################
def estCohort(plant):
	if 'Fall 2013' in plant.years:
		plant.updateCohort(1)
	elif 'Spring 2014' in plant.years:
		plant.updateCohort(2)
	elif 'Fall 2014' in plant.years:
		plant.updateCohort(3)
	elif 'Spring 2015' in plant.years:
		plant.updateCohort(4)

#####################################################################
# Formatting to make prints of plant objects human readable
#####################################################################
def info(season, year, site, row):
	return "Season: {}, Year: {}, Site: {}, Row: {}".format(season, year, site, row)

#####################################################################
# Check to see if there are any plants with very similar coordinates
# Anything that returns true will be considered the same plant
#####################################################################
def check(diff):
	if (-3 <= diff[0] <= 3) and (-3 <= diff[1] <= 3):
		return True
	else:
		return False

#####################################################################
# Sanitize string for leaves
#####################################################################
def sanitize_leaves(leaves):
	try:
		leaves = int(leaves)
	except:
		# leaves contains other characters
		if leaves == None or 'stalk' in leaves:
			print("Leaves changed to 6. Leaves was {}".format(leaves))
			return 6 # so it won't pass seedling check

		if 'coty' in leaves:
			print("Leaves changed to 0. Leaves was {}".format(leaves))
			return 0 # if there are still cotyledons, it's a seedling

		# For is usually: # alive, # dead
		# So we want to grab that first #
		new_string = ''
		for char in leaves:
			if char.isdigit():
				new_string += char
			if char == ' ':
				break

		if new_string == '':
			new_string = 0
			print ' * ' * 10
			print "NEW STRING WAS BLANK"
		print("Leaves changed to {}. Leaves was {}".format(new_string, leaves))
		leaves = int(new_string)
	return leaves

#####################################################################
# Check to see if the plant is a seedling
# If it is, consider it a new plant
#####################################################################
def checkSeedling(width, leaves):
	if width == None or float(width) <= 2.0:
		return True # It's a seedling

	# May neet to sanitize leaves
	try:
		leaves = int(leaves)
	except:
		# leaves contains other characters
		if leaves == None or 'stalk' in leaves:
			return False # if stalks are counted, it isn't a seedling

		if 'cot' in leaves:
			return True # if there are still cotyledons, it's a seedling

		# For is usually: # alive, # dead
		# So we want to grab that first #
		new_string = ''
		for char in leaves:
			if char.isdigit():
				new_string += char
			if char == ' ':
				break
		leaves = int(new_string)

	if leaves <= 5:
		return True # It's a seedling
	else:
		return False

#####################################################################
# Get some seedling data
#####################################################################
def seedlingData():

	global cohort1_seedling
	global cohort1_avg_life
	global cohort2_seedling
	global cohort2_avg_life
	global cohort3_seedling
	global cohort3_avg_life
	global cohort4_seedling
	global cohort4_avg_life
	global no_cohort_seedling

	global site1_seedling
	global site1_avg_life
	global site2_seedling
	global site2_avg_life
	global site3_seedling
	global site3_avg_life
	global site4_seedling
	global site4_avg_life

	global aspectN_seedling
	global aspectN_avg_life
	global aspectS_seedling
	global aspectS_avg_life

	global slopeA_seedling
	global slopeA_avg_life
	global slopeB_seedling
	global slopeB_avg_life
	global slopeC_seedling
	global slopeC_avg_life

	global total_plants
	global never_seedling # When we first found it, it didn't meet the criteria for seedling

	for plant in Individual.plants:
		estCohort(plant)

		if True not in plant.seedling.values():
			# If a plant was never considered a seedling, consider it a seedling the first year it was found.
			plant.addSeedlingYear(plant.years[0])

		total_plants += 1
		if plant.cohort == 1 and plant.seedling['Fall 2013'] == True:
			cohort1_seedling += 1
			cohort1_avg_life.append(len(plant.years))
		elif plant.cohort == 2 and plant.seedling['Spring 2014'] == True:
			cohort2_seedling += 1
			cohort2_avg_life.append(len(plant.years))
		elif plant.cohort == 3 and plant.seedling['Fall 2014'] == True:
			cohort3_seedling += 1
			cohort3_avg_life.append(len(plant.years))
		elif plant.cohort == 4 and plant.seedling['Spring 2015'] == True:
			cohort4_seedling += 1
			cohort4_avg_life.append(len(plant.years))
		elif plant.cohort == 99:
			no_cohort_seedling += 1
			print("This plant didn't have a cohort.")
			print plant
		else:
			print("This plant from cohort {} was never considered a seedling.".format(plant.cohort))
			print plant
			never_seedling += 1

		if plant.site == '1': # If the individual is at site 1, it was a seedling there
			site1_seedling += 1
			site1_avg_life.append(len(plant.years))
		elif plant.site == '2':
			site2_seedling += 1
			site2_avg_life.append(len(plant.years))
		elif plant.site == '3':
			site3_seedling += 1
			site3_avg_life.append(len(plant.years))
		elif plant.site == '4':
			site4_seedling += 1
			site4_avg_life.append(len(plant.years))
		else:
			print("!!!!!! SOMETHING WENT WRONG !!!!!!!!!") # This should never execute
			print plant

		if plant.aspect == 'N':
			aspectN_seedling += 1
			aspectN_avg_life.append(len(plant.years))
		elif plant.aspect == 'S':
			aspectS_seedling += 1
			aspectS_avg_life.append(len(plant.years))
		else:
			print("!!!!!! SOMETHING WENT WRONG !!!!!!!!!") # This should never execute
			print plant

		if plant.slope == 'A':
			slopeA_seedling += 1
			slopeA_avg_life.append(len(plant.years))
		elif plant.slope == 'B':
			slopeB_seedling += 1
			slopeB_avg_life.append(len(plant.years))
		elif plant.slope == 'C':
			slopeC_seedling += 1
			slopeC_avg_life.append(len(plant.years))
		else:
			print("!!!!!! SOMETHING WENT WRONG !!!!!!!!!") # This should never execute
			print plant

	# Just simplifying things here to make sure things are doing what I expect
	print("List of cohort 1 plants number of seasons alive")
	print cohort1_avg_life
	print("Sum of all seaons alive: {}".format(fsum(cohort1_avg_life)))
	print("Number of individuals making up that list: {}".format(len(cohort1_avg_life)))
	print("Average is sum divided by total individuals to get: {}".format(fsum(cohort1_avg_life)/len(cohort1_avg_life)))

	# Now on to the important stuff
	print("Number of seedlings in cohort 1: {}".format(cohort1_seedling))
	print("Average number of seasons cohort 1 seedlings lived: {}".format(fsum(cohort1_avg_life)/len(cohort1_avg_life)))
	print("Number of seedlings in cohort 2: {}".format(cohort2_seedling))
	print("Average number of seasons cohort 2 seedlings lived: {}".format(fsum(cohort2_avg_life)/len(cohort2_avg_life)))
	print("Number of seedlings in cohort 3: {}".format(cohort3_seedling))
	print("Average number of seasons cohort 3 seedlings lived: {}".format(fsum(cohort3_avg_life)/len(cohort3_avg_life)))
	print("Number of seedlings in cohort 4: {}".format(cohort4_seedling))
	print("Average number of seasons cohort 4 seedlings lived: {}".format(fsum(cohort4_avg_life)/len(cohort4_avg_life)))
	print("Number of seedlings not in a cohort: {}\n".format(no_cohort_seedling))

	print("Number of plants in site 1: {}".format(site1_seedling))
	print("Average number of seasons site 1 plants lived: {}".format(fsum(site1_avg_life)/len(site1_avg_life)))
	print("Number of plants in site 2: {}".format(site2_seedling))
	print("Average number of seasons site 2 plants lived: {}".format(fsum(site2_avg_life)/len(site2_avg_life)))
	print("Number of plants in site 3: {}".format(site3_seedling))
	print("Average number of seasons site 3 plants lived: {}".format(fsum(site3_avg_life)/len(site3_avg_life)))
	print("Number of plants in site 4: {}".format(site4_seedling))
	print("Average number of seasons site 4 plants lived: {}\n".format(fsum(site4_avg_life)/len(site4_avg_life)))

	print("Number of plants with North aspect: {}".format(aspectN_seedling))
	print("Average number of seasons North aspect plants lived: {}".format(fsum(aspectN_avg_life)/len(aspectN_avg_life)))
	print("Number of plants with South aspect: {}".format(aspectS_seedling))
	print("Average number of seasons South aspect plants lived: {}\n".format(fsum(aspectS_avg_life)/len(aspectS_avg_life)))

	print("Number of plants from slope position A: {}".format(slopeA_seedling))
	print("Average number of seasons slope position A plants lived: {}".format(fsum(slopeA_avg_life)/len(slopeA_avg_life)))
	print("Number of plants from slope position B: {}".format(slopeB_seedling))
	print("Average number of seasons slope position B plants lived: {}".format(fsum(slopeB_avg_life)/len(slopeB_avg_life)))
	print("Number of plants from slope position C: {}".format(slopeC_seedling))
	print("Average number of seasons slope position C plants lived: {}\n".format(fsum(slopeC_avg_life)/len(slopeC_avg_life)))

	print("{} total individuals.".format(total_plants))
	print("{} plants were never considered seedlings.".format(never_seedling))


#####################################################################
# Check to see if the plant was alive last season
# If it wasn't, it should be a new plant
#####################################################################
def lastSeason(plant, season, year):
	if season == 'Fall':
		previousSeason = 'Spring'
		previousYear = year
	else:
		previousSeason = 'Fall'
		previousYear = int(year) - 1
	if previousSeason + ' ' + str(previousYear) in plant.years:
		print 
		return True # It was around last season
	else:
		return False # It wasn't around last season

def getPreviousSeason(season, year):
	if season == 'Fall':
		previousSeason = 'Spring'
		previousYear = year
	else:
		previousSeason = 'Fall'
		previousYear = int(year) - 1
	return previousSeason + ' ' + str(previousYear)
############################################################################
# Goes through workbook based on given site, season and year
# Grabs attribute data to be added to plant list if it's a unique individual
############################################################################
def traverse(site, season, year):
	sheet = workbooks[season + year].get_sheet_by_name("Site " + site)
	for row in range(3, sheet.max_row + 1):
		#print(info(season, year, site, int(row)))
		if str(sheet['G' + str(row)].value).replace(".", "").isdigit():
			x_y = (int(sheet['G' + str(row)].value), int(sheet['H' + str(row)].value))
			aspect = sheet['C' + str(row)].value
			slope = sheet['D' + str(row)].value
			accession = sheet['E' + str(row)].value
			if sheet['I' + str(row)].value == None:
				condition = {season + ' ' + year : '-'}
			else:
				condition = {season + ' ' + year : sheet['I' + str(row)].value}
			height = {season + ' ' + year : sheet['K' + str(row)].value}
			width = {season + ' ' + year : sheet['L' + str(row)].value}
			leaves = {season + ' ' + year : sheet['M' + str(row)].value}
			seedling = None

			if (height == '-' and width == '-' and leaves == '-'):
				continue

			found = False
			for plant in Individual.plants:
				x_y_diff = tuple(numpy.subtract(x_y, plant.x_y))
				if (season + ' ' + year not in plant.years) and check(x_y_diff) and site == plant.site and aspect == plant.aspect and slope == plant.slope and accession == plant.accession and lastSeason(plant, season, year): 
					
					if 'D' not in condition[season + ' ' + year].upper() and checkSeedling(width[season + ' ' + year], leaves[season + ' ' + year]):
						print "*" * 20
						print "Seedling: {}, condition {}, year {}, {}, site {}, aspect {}, slope {}, accession {}, width {}, leaves {}".format(x_y, condition[season + ' ' + year], year, season, site, aspect, slope, accession, width, leaves)
						print ">>> This plant matches <<<"
						print plant

						clean_leaves = sanitize_leaves(leaves[season + ' ' + year])
						clean_plant_leaves = sanitize_leaves(plant.leaves[getPreviousSeason(season, year)])
						if plant.width[getPreviousSeason(season, year)] != None and float(width[season + ' ' + year]) < float(plant.width[getPreviousSeason(season, year)]) and clean_leaves < clean_plant_leaves:
							print "^ Seedling considered unique. New individual added. ^\n"
							seedling = season + ' ' + year
							Individual(site, x_y, aspect, slope, accession, season + " " + year, condition, height, width, leaves, seedling)						
							break

						print "^ Seedling NOT considered unique. Being combined with the above (and below, printed again). ^\n"

					found = True
					print "*********** Possible Conflict ***************"
					print "X, Y difference: " + str(x_y_diff)
					print "Individual at {}, condition {}, year {}, season {}, site {}, aspect {}, slope {}, accession {}, being added to: ".format(x_y, condition[season + ' ' + year], year, season, site, aspect, slope, accession)
					print plant
					print "*" * 20

					if 'D' not in condition[season + ' ' + year].upper(): 
						plant.addYear(season + " " + year)				# This way we can consider length of plant.years as number of seasons survived

					plant.addCondition(condition)
					plant.addHeight(height)
					plant.addWidth(width)
					plant.addLeaves(leaves)
					if seedling != None:
						plant.addSeedlingYear(seedling)
					break
			if not found:
				if checkSeedling(width[season + ' ' + year], leaves[season + ' ' + year]):
						seedling = season + ' ' + year
						print "Unique seedling: {} year {}, season {}, site {}, aspect {}, slope {}, accession {}".format(x_y, year, season, site, aspect, slope, accession)
				Individual(site, x_y, aspect, slope, accession, season + " " + year, condition, height, width, leaves, seedling)


#####################################################################
# Once we have all the plants in the list
# We use it to write it to a new workbook
#####################################################################
def writeOut():

	global cohort1_seedling
	global cohort1_avg_life
	global cohort2_seedling
	global cohort2_avg_life
	global cohort3_seedling
	global cohort3_avg_life
	global cohort4_seedling
	global cohort4_avg_life
	global no_cohort_seedling

	global site1_seedling
	global site1_avg_life
	global site2_seedling
	global site2_avg_life
	global site3_seedling
	global site3_avg_life
	global site4_seedling
	global site4_avg_life

	global aspectN_seedling
	global aspectN_avg_life
	global aspectS_seedling
	global aspectS_avg_life

	global slopeA_seedling
	global slopeA_avg_life
	global slopeB_seedling
	global slopeB_avg_life
	global slopeC_seedling
	global slopeC_avg_life

	global total_plants
	global never_seedling # When we first found it, it didn't meet the criteria for seedling

	wb = openpyxl.Workbook()
	sheet = wb.active
	ws_anal = wb.create_sheet()
	sheet['A1'] = 'Site'
	sheet['B1'] = 'Aspect'
	sheet['C1'] = 'Accession'
	sheet['D1'] = 'X (cm)'
	sheet['E1'] = 'Y (cm)'
	sheet['F1'] = 'Fall 2013'
	sheet['G1'] = 'Spring 2014'
	sheet['H1'] = 'Fall 2014'
	sheet['I1'] = 'Spring 2015'
	sheet['J1'] = 'Fall 2015'
	sheet['L2'].fill = cohort1Color
	sheet['L2'] = 'Cohort 1'
	sheet['L3'].fill = cohort2Color
	sheet['L3'] = 'Cohort 2'
	sheet['L4'].fill = cohort3Color
	sheet['L4'] = 'Cohort 3'
	sheet['L5'].fill = cohort4Color
	sheet['L5'] = 'Cohort 4'

	ws_anal['A1'] = 'Total Plants'
	ws_anal['A2'] = total_plants
	ws_anal['B1'] = 'Total number of seedlings in cohort 1'
	ws_anal['B2'] = cohort1_seedling
	ws_anal['C1'] = 'Total number of seedlings in cohort 2'
	ws_anal['C2'] = cohort2_seedling
	ws_anal['D1'] = 'Total number of seedlings in cohort 3'
	ws_anal['D2'] = cohort3_seedling
	ws_anal['E1'] = 'Total number of seedlings in cohort 4'
	ws_anal['E2'] = cohort4_seedling
	ws_anal['F1'] = 'Total number of plants not in a cohort'
	ws_anal['F2'] = no_cohort_seedling
	ws_anal['B4'] = 'Average number of seasons cohort 1 seedlings lived'
	ws_anal['B5'] = fsum(cohort1_avg_life)/len(cohort1_avg_life)
	ws_anal['C4'] = 'Average number of seasons cohort 2 seedlings lived'
	ws_anal['C5'] = fsum(cohort2_avg_life)/len(cohort2_avg_life)
	ws_anal['D4'] = 'Average number of seasons cohort 3 seedlings lived'
	ws_anal['D5'] = fsum(cohort3_avg_life)/len(cohort3_avg_life)
	ws_anal['E4'] = 'Average number of seasons cohort 4 seedlings lived'
	ws_anal['E5'] = fsum(cohort4_avg_life)/len(cohort4_avg_life)
	ws_anal['B7'] = 'Number of plants at site 1'
	ws_anal['B8'] = site1_seedling
	ws_anal['C7'] = 'Number of plants at site 2'
	ws_anal['C8'] = site2_seedling
	ws_anal['D7'] = 'Number of plants at site 3'
	ws_anal['D8'] = site3_seedling
	ws_anal['E7'] = 'Number of plants at site 4'
	ws_anal['E8'] = site4_seedling
	ws_anal['B10'] = 'Average number of seasons site 1 plants lived'
	ws_anal['B11'] = fsum(site1_avg_life)/len(site1_avg_life)
	ws_anal['C10'] = 'Average number of seasons site 2 plants lived'
	ws_anal['C11'] = fsum(site2_avg_life)/len(site2_avg_life)
	ws_anal['D10'] = 'Average number of seasons site 3 plants lived'
	ws_anal['D11'] = fsum(site3_avg_life)/len(site3_avg_life)
	ws_anal['E10'] = 'Average number of seasons site 4 plants lived'
	ws_anal['E11'] = fsum(site4_avg_life)/len(site4_avg_life)
	ws_anal['B13'] = 'Number of plants at north aspect'
	ws_anal['B14'] = aspectN_seedling
	ws_anal['C13'] = 'Number of plants at south aspect'
	ws_anal['C14'] = aspectS_seedling
	ws_anal['B16'] = 'Average number of seasons north aspect plants lived'
	ws_anal['B17'] = fsum(aspectN_avg_life)/len(aspectN_avg_life)
	ws_anal['C16'] = 'Average number of seasons south aspect plants lived'
	ws_anal['C17'] = fsum(aspectS_avg_life)/len(aspectS_avg_life)
	ws_anal['B19'] = 'Number of plants at slope position A'
	ws_anal['B20'] = slopeA_seedling
	ws_anal['C19'] = 'Number of plants at slope position B'
	ws_anal['C20'] = slopeB_seedling
	ws_anal['D19'] = 'Number of plants at slope position C'
	ws_anal['D20'] = slopeC_seedling
	ws_anal['B22'] = 'Average number of seasons slope position A plants lived'
	ws_anal['B23'] = fsum(slopeA_avg_life)/len(slopeA_avg_life)
	ws_anal['C22'] = 'Average number of seasons slope position B plants lived'
	ws_anal['C23'] = fsum(slopeB_avg_life)/len(slopeB_avg_life)
	ws_anal['D22'] = 'Average number of seasons slope position C plants lived'
	ws_anal['D23'] = fsum(slopeC_avg_life)/len(slopeC_avg_life)

	for idx, plant in enumerate(Individual.plants):
		color = estColor(plant)
		sheet['A' + str(idx + 2)] = plant.site
		sheet['B' + str(idx + 2)] = plant.aspect
		sheet['C' + str(idx + 2)] = plant.accession
		sheet['D' + str(idx + 2)] = plant.x_y[0]
		if color is not None:
			sheet['D' + str(idx + 2)].fill = color
		sheet['E' + str(idx + 2)] = plant.x_y[1]
		if color is not None:
			sheet['E' + str(idx + 2)].fill = color

		if 'Fall 2013' in plant.condition:
			sheet['F' + str(idx + 2)] = plant.condition['Fall 2013']
		else:
			sheet['F' + str(idx + 2)] = '-'

		if 'Spring 2014' in plant.condition:
			sheet['G' + str(idx + 2)] = plant.condition['Spring 2014']
		else:
			sheet['G' + str(idx + 2)] = '-'

		if 'Fall 2014' in plant.condition:
			sheet['H' + str(idx + 2)] = plant.condition['Fall 2014']
		else:
			sheet['H' + str(idx + 2)] = '-'

		if 'Spring 2015' in plant.condition:
			sheet['I' + str(idx + 2)] = plant.condition['Spring 2015']
		else:
			sheet['I' + str(idx + 2)] = '-'

		if 'Fall 2015' in plant.condition:
			sheet['J' + str(idx + 2)] = plant.condition['Fall 2015']
		else:
			sheet['J' + str(idx + 2)] = '-'

	wb.save('2013-2015_Survivorship_Table_Final_seedlings.xlsx')

######################################################################
# For Fall 2013
# For Site 1 Fall 2013
fall_sheet1 = wb_Fall_2013.get_sheet_by_name('Site 1')
traverse('1', 'Fall', '2013')

# For Site 2 Fall 2013
traverse('2', 'Fall', '2013')		

# For Site 3 Fall 2013
traverse('3', 'Fall', '2013')

# For Site 4 Fall 2013
traverse('4', 'Fall', '2013')

#########################################################################
# For Spring 2014
# For Site 1 Spring 2014
traverse('1', 'Spring', '2014')

# For Site 2 Spring 2014
traverse('2', 'Spring', '2014')

# For Site 3 Spring 2014
traverse('3', 'Spring', '2014')

# For Site 4 Spring 2014
traverse('4', 'Spring', '2014')

#########################################################################
# For Fall 2014
# For Site 1 Fall 2014
traverse('1', 'Fall', '2014')

# For Site 2 Fall 2014
traverse('2', 'Fall', '2014')

# For Site 3 Fall 2014
traverse('3', 'Fall', '2014')

# For Site 4 Fall 2014
traverse('4', 'Fall', '2014')

#########################################################################
# For Spring 2015
# For Site 1 Spring 2015
traverse('1', 'Spring', '2015')

# For Site 2 Spring 2015
traverse('2', 'Spring', '2015')

# For Site 3 Spring 2015
traverse('3', 'Spring', '2015')

# For Site 4 Spring 2015
traverse('4', 'Spring', '2015')

#########################################################################
# For Fall 2015
# For Site 1 Fall 2015
traverse('1', 'Fall', '2015')

# For Site 2 Fall 2015
traverse('2', 'Fall', '2015')

# For Site 3 Fall 2015
traverse('3', 'Fall', '2015')

# For Site 4 Fall 2015
traverse('4', 'Fall', '2015')

#########################################################################
# Additional Commands

seedlingData()
writeOut()

'''
for idx, plant in enumerate(Individual.plants):
	x_y_diff = tuple(numpy.subtract((41, 20), plant.x_y))
	if check(x_y_diff) and '2' == plant.site and 'S' == plant.aspect and 'C' == plant.slope and 'S-448' == plant.accession:
		print plant
'''